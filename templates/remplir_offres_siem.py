#!/usr/bin/env python3
"""Remplit le template Excel avec l'évaluation Sentinel / FortiSIEM / Wazuh."""

from copy import copy
from pathlib import Path

from openpyxl import load_workbook

SRC = Path(__file__).resolve().parent / "Evaluation_SIEM_NAC_3_offres.xlsx"
OUT = Path(__file__).resolve().parent / "Evaluation_SIEM_NAC_3_offres_REMPLI.xlsx"

# --- Paramètres projet (adapter ici) ---
BUDGET_FCFA = 50_000_000  # TCO cible 3 ans
EUR_PER_FCFA = 1 / 655.957  # XOF (Franc CFA BCEAO)
BUDGET_EUR_REF = round(BUDGET_FCFA * EUR_PER_FCFA)  # ~76 220 € HT indicatif
PRIORITY_OFFER = "A"  # Offre prioritaire : Microsoft Sentinel
# (conformité, note, commentaire) pour A=Sentinel, B=FortiSIEM, C=Wazuh
SIEM_DATA = {
    5: (
        ("Oui", 5, "Connecteurs cloud natifs, AMA, API, syslog via CEF/AMA"),
        ("Oui", 4, "Collecte multi-vendeurs via agents FortiSIEM / collectors"),
        ("Oui", 4, "Agents Wazuh + syslog + modules cloud (AWS/Azure/GCP)"),
    ),
    6: (
        ("Oui", 5, "Règles analytics KQL, scheduled queries, NRT"),
        ("Oui", 4, "Règles temps réel + corrélation multi-étapes"),
        ("Oui", 4, "Règles custom YAML, decoders, corrélation intégrée"),
    ),
    7: (
        ("Partiel", 4, "UEBA via add-on Sentinel / intégration Defender XDR"),
        ("Partiel", 3, "Analyse comportementale limitée vs leaders UEBA"),
        ("Partiel", 2, "Pas d'UEBA natif ; détection par règles et FIM"),
    ),
    8: (
        ("Oui", 5, "Microsoft TI, STIX/TAXII, enrichissement Defender"),
        ("Oui", 4, "FortiGuard TI intégré"),
        ("Partiel", 3, "Intégration TI possible (MISP, VirusTotal, etc.)"),
    ),
    9: (
        ("Oui", 5, "Playbooks Sentinel + Logic Apps / Defender automation"),
        ("Partiel", 3, "Automatisation intégrée ; FortiSOAR en complément"),
        ("Partiel", 2, "Active Response basique ; pas de SOAR complet natif"),
    ),
    10: (
        ("Oui", 5, "Solution gallery, templates Microsoft (MITRE, cloud, identity)"),
        ("Oui", 4, "Use cases pré-packagés secteur / compliance"),
        ("Oui", 4, "Templates PCI-DSS, CIS, GDPR, NIST"),
    ),
    11: (
        ("Oui", 5, "Hunting KQL, notebooks, Defender advanced hunting"),
        ("Oui", 3, "FortiView / recherche ; moins riche que KQL"),
        ("Oui", 4, "Wazuh indexer + OpenSearch Dashboards / API"),
    ),
    12: (
        ("Oui", 4, "Workbooks, Power BI, rapports conformité Azure"),
        ("Oui", 4, "Dashboards opérationnels + reporting compliance"),
        ("Partiel", 3, "Dashboards OSS ; personnalisation technique requise"),
    ),
    14: (
        ("Oui", 5, "Scalabilité cloud Azure (ingestion massive)"),
        ("Oui", 4, "Scalable en cluster collectors / appliances"),
        ("Partiel", 3, "Scalable si architecture indexer/cluster dimensionnée"),
    ),
    15: (
        ("Oui", 5, "SLA Azure, redondance régionale"),
        ("Oui", 4, "HA collectors + appliances en mode cluster"),
        ("Partiel", 3, "HA possible (cluster) mais à architecturer"),
    ),
    16: (
        ("Oui", 4, "Rétention ADX / Archive ; coût long terme"),
        ("Oui", 4, "Rétention on-prem configurable par politique"),
        ("Partiel", 3, "Rétention liée au stockage OpenSearch / ILM"),
    ),
    17: (
        ("Oui", 4, "Alerting NRT ; latence dépend ingestion pipeline"),
        ("Oui", 4, "Corrélation temps réel performante"),
        ("Partiel", 3, "Bonne latence à l'échelle si cluster bien dimensionné"),
    ),
    18: (
        ("Oui", 5, "SaaS cloud Azure (hybride via AMA on-prem)"),
        ("Oui", 4, "On-prem / hybride / cloud collectors"),
        ("Oui", 5, "On-prem / cloud (Wazuh Cloud option) — open source"),
    ),
    20: (
        ("Oui", 5, "Intégration native Entra ID / Azure AD, Identity Protection"),
        ("Oui", 4, "Intégration AD / LDAP pour corrélation"),
        ("Partiel", 3, "Intégration AD via agents / API ; moins native"),
    ),
    21: (
        ("Oui", 5, "Defender XDR / EDR intégration native"),
        ("Oui", 5, "FortiEDR / FortiClient intégration forte"),
        ("Partiel", 3, "Agents Wazuh (FIM, vuln) ; pas EDR commercial complet"),
    ),
    22: (
        ("Oui", 4, "Connecteurs Palo Alto, Fortinet, Check Point, etc."),
        ("Oui", 5, "FortiGate intégration native excellente"),
        ("Oui", 4, "Syslog / intégrations firewall standards"),
    ),
    23: (
        ("Oui", 4, "Logic Apps, ServiceNow, Jira via connecteurs"),
        ("Partiel", 3, "Connecteurs ticketing disponibles mais limités"),
        ("Partiel", 3, "Intégrations webhook / scripts custom"),
    ),
    24: (
        ("Partiel", 3, "Via syslog générique ; pas couplage NAC natif"),
        ("Oui", 5, "Intégration FortiNAC native dans l'écosystème Fortinet"),
        ("Partiel", 2, "Réception événements NAC via syslog uniquement"),
    ),
    25: (
        ("Oui", 5, "M365, Azure natif ; connecteurs AWS/GCP"),
        ("Partiel", 3, "Connecteurs cloud présents mais moins exhaustifs"),
        ("Oui", 4, "Modules cloud AWS, Azure, GCP, Office 365"),
    ),
    27: (
        ("Partiel", 3, "Courbe KQL / Azure ; puissant mais exigeant"),
        ("Oui", 4, "Console unifiée Fortinet familière aux équipes réseau"),
        ("Partiel", 3, "Interface technique ; expertise OSS recommandée"),
    ),
    28: (
        ("Partiel", 3, "Tuning analytics rules nécessaire (faux positifs)"),
        ("Partiel", 3, "Tuning règles requis selon périmètre"),
        ("Partiel", 3, "Tuning règles / decoders indispensable"),
    ),
    29: (
        ("Oui", 5, "Azure RBAC, workspaces multi-équipes"),
        ("Oui", 4, "RBAC multi-organisation / multi-site"),
        ("Oui", 4, "RBAC Wazuh + multi-tenant possible"),
    ),
    30: (
        ("Partiel", 3, "Charge admin + gestion coûts ingestion"),
        ("Oui", 4, "Administration centralisée si stack Fortinet"),
        ("Partiel", 2, "Charge ops élevée (patch, cluster, règles)"),
    ),
    31: (
        ("Oui", 5, "Documentation Microsoft extensive + GitHub"),
        ("Oui", 4, "Documentation Fortinet + guides use cases"),
        ("Oui", 4, "Documentation Wazuh + communauté active"),
    ),
    33: (
        ("Oui", 5, "Chiffrement Azure (TLS, stockage chiffré)"),
        ("Oui", 4, "Chiffrement transit / repos selon déploiement"),
        ("Oui", 4, "TLS, chiffrement configurable OpenSearch"),
    ),
    34: (
        ("Partiel", 3, "Données dans région Azure choisie ; cloud US possible"),
        ("Oui", 4, "On-prem = souveraineté maximale"),
        ("Oui", 5, "100 % on-prem possible ; contrôle total des données"),
    ),
    35: (
        ("Oui", 5, "Azure Activity Log / audit intégré"),
        ("Oui", 4, "Journalisation admin appliances"),
        ("Oui", 4, "Audit des actions administrateurs"),
    ),
    36: (
        ("Oui", 5, "ISO 27001, SOC2, RGPD, certifications Azure"),
        ("Oui", 4, "Conformité sectorielle via reporting intégré"),
        ("Oui", 4, "Modules compliance PCI, HIPAA, CIS, GDPR"),
    ),
    38: (
        ("Oui", 4, "Support Microsoft entreprise (selon contrat)"),
        ("Oui", 4, "Support Fortinet 24/7 (selon contrat FortiCare)"),
        ("Partiel", 3, "Support communauté ; support commercial Wazuh en option"),
    ),
    39: (
        ("Oui", 4, "Formations Microsoft Learn / partenaires"),
        ("Oui", 4, "Formation Fortinet NSE disponible"),
        ("Partiel", 3, "Formation disponible ; montée en compétence interne"),
    ),
    40: (
        ("Oui", 5, "Roadmap Microsoft Security forte / pérennité"),
        ("Oui", 4, "Éditeur pérenne, roadmap intégration stack"),
        ("Oui", 4, "Projet OSS actif, adoption croissante"),
    ),
    41: (
        ("Oui", 5, "Références enterprise / secteur public mondial"),
        ("Oui", 4, "Références mid-market / entreprises stack Fortinet"),
        ("Partiel", 3, "Références croissantes ; moins enterprise historique"),
    ),
    43: (
        ("Partiel", 2, "Modèle ingestion GB/jour + coûts variables"),
        ("Partiel", 3, "Licence par EPS/devices — modèle compréhensible"),
        ("Oui", 5, "Open source + support optionnel — très clair"),
    ),
    44: (
        ("Oui", 5, f"Budget 50 M FCFA / 3 ans — dimensionnement aligné"),
        ("Partiel", 2, "Hors budget 50 M FCFA (stack SIEM+NAC)"),
        ("Oui", 4, "TCO bas mais hors priorité stratégique"),
    ),
    45: (
        ("Partiel", 4, "Ingestion maîtrisée (filtrage AMA) — surveiller burst"),
        ("Partiel", 2, "Licences multiples + appliances hors enveloppe"),
        ("Partiel", 3, "Coût infra + RH à anticiper"),
    ),
}

# FortiNAC pour offre B uniquement ; A/C = SIEM seul
NAC_DATA = {
    5: (
        ("N/A", 0, "NAC non proposé dans l'offre A (SIEM seul)"),
        ("Oui", 5, "802.1X filaire et Wi-Fi natif FortiNAC"),
        ("N/A", 0, "NAC non proposé (Wazuh = SIEM/XDR open source)"),
    ),
    6: (
        ("N/A", 0, ""),
        ("Oui", 4, "MAB pour équipements non 802.1X"),
        ("N/A", 0, ""),
    ),
    7: (
        ("N/A", 0, ""),
        ("Oui", 5, "Portail captive + sponsoring invités"),
        ("N/A", 0, ""),
    ),
    8: (
        ("N/A", 0, ""),
        ("Oui", 5, "Politiques dynamiques VLAN / ACL / quarantine"),
        ("N/A", 0, ""),
    ),
    9: (
        ("N/A", 0, ""),
        ("Oui", 4, "Posture check (AV, patch) intégré"),
        ("N/A", 0, ""),
    ),
    10: (
        ("N/A", 0, ""),
        ("Oui", 4, "BYOD, employés, partenaires"),
        ("N/A", 0, ""),
    ),
    11: (
        ("N/A", 0, ""),
        ("Oui", 4, "Profilage IoT / OT / devices non managés"),
        ("N/A", 0, ""),
    ),
    13: (
        ("N/A", 0, ""),
        ("Oui", 4, "Découverte et inventaire réseau"),
        ("N/A", 0, ""),
    ),
    14: (
        ("N/A", 0, ""),
        ("Oui", 4, "Fingerprint actif / passif"),
        ("N/A", 0, ""),
    ),
    15: (
        ("N/A", 0, ""),
        ("Oui", 4, "Classification automatique des devices"),
        ("N/A", 0, ""),
    ),
    16: (
        ("N/A", 0, ""),
        ("Oui", 4, "Visibilité temps réel endpoint / switch / port"),
        ("N/A", 0, ""),
    ),
    18: (
        ("N/A", 0, ""),
        ("Oui", 4, "Multi-vendeurs switches (Cisco, Aruba, etc.)"),
        ("N/A", 0, ""),
    ),
    19: (
        ("N/A", 0, ""),
        ("Oui", 5, "Intégration FortiGate WLC / contrôleurs Wi-Fi"),
        ("N/A", 0, ""),
    ),
    20: (
        ("N/A", 0, ""),
        ("Oui", 4, "VPN / accès distant"),
        ("N/A", 0, ""),
    ),
    21: (
        ("N/A", 0, ""),
        ("Oui", 5, "AD / Entra ID / RADIUS / PKI"),
        ("N/A", 0, ""),
    ),
    22: (
        ("N/A", 0, ""),
        ("Partiel", 3, "Intégration MDM via API / posture"),
        ("N/A", 0, ""),
    ),
    23: (
        ("N/A", 0, ""),
        ("Oui", 5, "Intégration FortiSIEM + FortiGate"),
        ("N/A", 0, ""),
    ),
    25: (
        ("N/A", 0, ""),
        ("Oui", 5, "Mode audit puis enforce sans coupure"),
        ("N/A", 0, ""),
    ),
    26: (
        ("N/A", 0, ""),
        ("Oui", 4, "Déploiement agents / appliances standard Fortinet"),
        ("N/A", 0, ""),
    ),
    27: (
        ("N/A", 0, ""),
        ("Oui", 4, "HA policy server en cluster"),
        ("N/A", 0, ""),
    ),
    28: (
        ("N/A", 0, ""),
        ("Oui", 4, "Multi-sites / fédération"),
        ("N/A", 0, ""),
    ),
    30: (
        ("N/A", 0, ""),
        ("Partiel", 3, "Micro-segmentation via intégration FortiGate"),
        ("N/A", 0, ""),
    ),
    31: (
        ("N/A", 0, ""),
        ("Oui", 4, "Quarantine automatisée"),
        ("N/A", 0, ""),
    ),
    32: (
        ("N/A", 0, ""),
        ("Oui", 4, "Reporting conformité accès réseau"),
        ("N/A", 0, ""),
    ),
    34: (
        ("N/A", 0, ""),
        ("Oui", 4, "SLA FortiCare"),
        ("N/A", 0, ""),
    ),
    35: (
        ("N/A", 0, ""),
        ("Oui", 4, "Formation Fortinet NSE"),
        ("N/A", 0, ""),
    ),
    36: (
        ("N/A", 0, ""),
        ("Oui", 4, "Roadmap Fortinet NAC"),
        ("N/A", 0, ""),
    ),
    37: (
        ("N/A", 0, ""),
        ("Oui", 4, "Références clients FortiNAC"),
        ("N/A", 0, ""),
    ),
    39: (
        ("N/A", 0, ""),
        ("Oui", 4, "Licence par endpoint claire"),
        ("N/A", 0, ""),
    ),
    40: (
        ("N/A", 0, ""),
        ("Oui", 4, "Bon alignement stack Fortinet"),
        ("N/A", 0, ""),
    ),
    41: (
        ("N/A", 0, ""),
        ("Partiel", 3, "Appliances + renew"),
        ("N/A", 0, ""),
    ),
}


def fill_criteria(ws, data: dict):
    for row, triples in data.items():
        for i, (conf, note, comment) in enumerate(triples):
            base = 3 + i * 3
            ws.cell(row=row, column=base, value=conf)
            ws.cell(row=row, column=base + 1, value=note)
            ws.cell(row=row, column=base + 2, value=comment)


def main():
    wb = load_workbook(SRC)

    ctx = wb["02_Contexte"]
    # Identification offres
    ctx["B5"] = "Microsoft Sentinel"
    ctx["C5"] = "FortiSIEM"
    ctx["D5"] = "Wazuh"
    ctx["B6"] = "Non proposé (offre SIEM seule)"
    ctx["C6"] = "FortiNAC"
    ctx["D6"] = "Non proposé (offre SIEM seule)"
    ctx["B7"] = "Microsoft / partenaire"
    ctx["C7"] = "Fortinet / intégrateur"
    ctx["D7"] = "Wazuh / intégrateur OSS"
    ctx["B8"] = "Éval. agent — priorité Offre A"
    ctx["C8"] = "Référence comparative"
    ctx["D8"] = "Référence comparative"
    ctx["B9"] = "06/08/2026"
    ctx["C9"] = "06/08/2026"
    ctx["D9"] = "06/08/2026"
    ctx["B10"] = "À confirmer avec vendeur"
    ctx["C10"] = "À confirmer avec vendeur"
    ctx["D10"] = "À confirmer avec vendeur"
    ctx["B11"] = "—"
    ctx["C11"] = "—"
    ctx["D11"] = "—"
    ctx["B12"] = "SaaS (Azure) / hybride"
    ctx["C12"] = "On-prem / hybride"
    ctx["D12"] = "On-prem / cloud Wazuh"

    # Périmètre calibré pour budget 50 M FCFA / 3 ans (Sentinel)
    ctx["B16"] = 450
    ctx["E16"] = "Calibré budget 50 M FCFA"
    ctx["B17"] = 100
    ctx["E17"] = "~100 EPS — sources prioritaires"
    ctx["B18"] = 25
    ctx["E18"] = "Go/jour — avec filtrage ingestion"
    ctx["B19"] = 60
    ctx["E19"] = "Rétention chaude Azure"
    ctx["B20"] = 12
    ctx["E20"] = "Archive basique"
    ctx["B21"] = 2
    ctx["E21"] = "Sites principaux"
    ctx["B22"] = 120
    ctx["E22"] = "NAC phase 2 (hors budget initial)"
    ctx["B23"] = "Filaire + Wi-Fi + M365 / Entra ID"
    ctx["B24"] = "Oui"
    ctx["B25"] = "Partiel"
    ctx["E25"] = "HA Azure région France / EU"
    ctx["B26"] = "Région Azure EU — données cloud"
    ctx["B27"] = "SOC interne — montée en compétences KQL"
    ctx["B28"] = BUDGET_FCFA
    ctx["C28"] = "FCFA"
    ctx["E28"] = (
        f"TCO 3 ans max | réf. ~{BUDGET_EUR_REF:,} € HT "
        f"(1 € = 655,957 FCFA) — montants TCO en FCFA"
    ).replace(",", " ")
    ctx["B29"] = "Q2 2027"
    ctx["E29"] = "Déploiement progressif Sentinel"

    # Pondérations SIEM — priorité critères Sentinel (cloud, intégrations, détection)
    ctx["B34"] = 30   # Fonctionnalités & détection
    ctx["B35"] = 15   # Architecture
    ctx["B36"] = 25   # Intégrations (M365 / Entra / cloud)
    ctx["B37"] = 10   # Exploitation SOC
    ctx["B38"] = 10   # Sécurité / conformité
    ctx["B39"] = 5    # Support
    ctx["B40"] = 5    # TCO (budget contraint — déjà cadré)

    # Pondérations NAC — phase 2 hors enveloppe initiale
    ctx["B46"] = 20
    ctx["B47"] = 15
    ctx["B48"] = 20
    ctx["B49"] = 15
    ctx["B50"] = 15
    ctx["B51"] = 10
    ctx["B52"] = 10

    # Projet : SIEM prioritaire (NAC reporté)
    ctx["B57"] = 90
    ctx["B58"] = 10

    fill_criteria(wb["03_SIEM_Critères"], SIEM_DATA)
    fill_criteria(wb["04_NAC_Critères"], NAC_DATA)

    # TCO 3 ans en FCFA — Offre A calée sur 50 M ; B/C hors budget (référence)
    tco = wb["05_TCO"]
    tco["A1"] = "COÛT TOTAL DE POSSESSION (TCO) — 3 ANS — FCFA (Franc CFA)"
    costs = {
        # Offre A (Sentinel) — total ~50 M FCFA
        5: (10_500_000, 28_000_000, 12_000_000),
        6: (11_000_000, 29_000_000, 12_500_000),
        7: (11_500_000, 30_000_000, 13_000_000),
        8: (1_500_000, 5_000_000, 0),
        9: (1_000_000, 3_000_000, 1_000_000),
        10: (500_000, 2_000_000, 500_000),
        12: (0, 12_000_000, 0),
        13: (0, 12_500_000, 0),
        14: (0, 13_000_000, 0),
        15: (0, 3_000_000, 0),
        16: (0, 8_000_000, 0),
        17: (0, 2_000_000, 0),
        19: (2_500_000, 8_000_000, 4_000_000),
        20: (0, 6_000_000, 0),
        21: (1_500_000, 4_000_000, 1_500_000),
        22: (1_000_000, 3_000_000, 1_000_000),
        23: (1_500_000, 4_000_000, 1_500_000),
        25: (1_500_000, 2_000_000, 1_000_000),
        26: (1_000_000, 1_500_000, 500_000),
        27: (500_000, 1_000_000, 0),
        29: (0, 0, 0),  # support inclus Azure / FortiCare
        30: (0, 0, 0),
        31: (0, 0, 0),
        32: (0, 0, 0),
        34: (0, 0, 3_000_000),
        35: (0, 0, 500_000),
        36: (1_000_000, 5_000_000, 1_500_000),
    }
    tco_comment = (
        f"FCFA — Budget max {BUDGET_FCFA:,} — Offre A prioritaire"
    ).replace(",", " ")
    for row, (a, b, c) in costs.items():
        tco.cell(row=row, column=2, value=a)
        tco.cell(row=row, column=3, value=b)
        tco.cell(row=row, column=4, value=c)
        tco.cell(row=row, column=5, value="FCFA")
        tco.cell(row=row, column=6, value=tco_comment)

    syn = wb["07_Synthese"]
    syn["B5"] = "Microsoft Sentinel"
    syn["C5"] = "FortiSIEM + FortiNAC"
    syn["D5"] = "Wazuh"
    syn["B16"] = (
        "PRIORITÉ STRATÉGIQUE. Écosystème Microsoft (M365, Entra, Defender). "
        "Détection KQL, SOAR Logic Apps, scalabilité cloud. "
        "Dimensionné pour budget 50 M FCFA / 3 ans."
    )
    syn["C16"] = (
        "Stack Fortinet complète SIEM+NAC — référence comparative. "
        "Hors enveloppe budgétaire (~95 M FCFA estimés)."
    )
    syn["D16"] = (
        "Alternative OSS souveraine — TCO licence faible (~38 M FCFA) "
        "mais non prioritaire vs Sentinel."
    )
    syn["B17"] = (
        "NAC non inclus phase 1 ; courbe KQL à planifier. "
        "Ingestion à monitorer (filtrage AMA)."
    )
    syn["C17"] = "Budget 50 M FCFA insuffisant pour SIEM+NAC Fortinet"
    syn["D17"] = "Charge ops élevée ; écart fonctionnel vs Sentinel cloud"
    syn["B18"] = "Dépendance Azure — mitigée par priorité Offre A"
    syn["C18"] = "TCO ~2× budget cible"
    syn["D18"] = "Moins d'intégrations M365 natives"
    syn["B19"] = "Confirmer région Azure EU + engagement ingestion"
    syn["C19"] = "Non sélectionné — hors budget"
    syn["D19"] = "Plan B technique uniquement"
    syn["B21"] = "Offre A — Microsoft Sentinel (prioritaire)"
    syn["A23"] = (
        "L'offre A (Microsoft Sentinel) est recommandée en priorité : elle respecte "
        f"l'enveloppe budgétaire de 50 000 000 FCFA sur 3 ans (périmètre calibré : "
        "~100 EPS, ~450 utilisateurs, 2 sites), maximise l'intégration M365/Entra ID "
        "et offre le meilleur niveau de détection/SOAR dans cette enveloppe. "
        "Les offres B (~95 M FCFA) et C (~38 M FCFA) servent de référence : B excède "
        "le budget ; C est moins coûteuse mais ne répond pas à la priorité stratégique "
        "Microsoft. Le NAC est reporté en phase 2 (budget séparé ou 802.1X progressif)."
    )
    syn["A29"] = (
        "Plan B : Wazuh si contrainte souveraineté absolue. "
        "NAC phase 2 : solution tiers ou extension budget."
    )
    syn["A33"] = (
        "1. PoC Sentinel 30 jours (sources M365 + AD + firewall) | "
        "2. Atelier ingestion / filtrage AMA pour tenir le budget | "
        "3. Négociation EES / Azure commit Microsoft | "
        "4. Formation KQL SOC (5 jours) | 5. Roadmap NAC phase 2"
    )

    wb.save(OUT)
    print(f"Fichier rempli : {OUT}")


if __name__ == "__main__":
    main()
