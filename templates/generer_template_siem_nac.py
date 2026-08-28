#!/usr/bin/env python3
"""Génère le template Excel d'évaluation comparative SIEM + NAC (3 offres)."""

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.formatting.rule import FormulaRule, ColorScaleRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from pathlib import Path

OUT = Path(__file__).resolve().parent / "Evaluation_SIEM_NAC_3_offres.xlsx"

# --- Styles ---
THIN = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)
FILL_HEADER = PatternFill("solid", fgColor="1F4E79")
FILL_SECTION = PatternFill("solid", fgColor="2E75B6")
FILL_SUB = PatternFill("solid", fgColor="D6E3F0")
FILL_INPUT = PatternFill("solid", fgColor="FFF2CC")
FILL_CALC = PatternFill("solid", fgColor="E2EFDA")
FILL_TITLE = PatternFill("solid", fgColor="0D3B66")
FILL_WARN = PatternFill("solid", fgColor="FCE4D6")
FONT_WHITE = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
FONT_TITLE = Font(name="Calibri", bold=True, color="FFFFFF", size=16)
FONT_SECTION = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
FONT_BOLD = Font(name="Calibri", bold=True, size=11)
FONT_NORMAL = Font(name="Calibri", size=10)
FONT_SMALL = Font(name="Calibri", size=9, italic=True, color="666666")
ALIGN_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_L = Alignment(horizontal="left", vertical="center", wrap_text=True)


def style_header_row(ws, row, start_col, end_col):
    for c in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = FILL_HEADER
        cell.font = FONT_WHITE
        cell.alignment = ALIGN_C
        cell.border = THIN


def style_section(ws, row, start_col, end_col, label):
    ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
    cell = ws.cell(row=row, column=start_col, value=label)
    cell.fill = FILL_SECTION
    cell.font = FONT_SECTION
    cell.alignment = ALIGN_L
    for c in range(start_col, end_col + 1):
        ws.cell(row=row, column=c).fill = FILL_SECTION
        ws.cell(row=row, column=c).border = THIN
        ws.cell(row=row, column=c).font = FONT_SECTION


def style_input(cell):
    cell.fill = FILL_INPUT
    cell.border = THIN
    cell.alignment = ALIGN_C
    cell.font = FONT_NORMAL


def style_calc(cell):
    cell.fill = FILL_CALC
    cell.border = THIN
    cell.alignment = ALIGN_C
    cell.font = FONT_BOLD


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_score_validation(ws, *ranges):
    """Note 0-5."""
    dv = DataValidation(
        type="list",
        formula1='"0,1,2,3,4,5"',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Note invalide",
        error="Saisir une note de 0 à 5.",
    )
    for rng in ranges:
        dv.add(rng)
    ws.add_data_validation(dv)


def add_oui_non(ws, *ranges):
    dv = DataValidation(
        type="list",
        formula1='"Oui,Partiel,Non,N/A"',
        allow_blank=True,
    )
    for rng in ranges:
        dv.add(rng)
    ws.add_data_validation(dv)


# =============================================================================
# Feuille 1 — Instructions
# =============================================================================
def build_instructions(wb):
    ws = wb.active
    ws.title = "01_Instructions"
    set_col_widths(ws, [8, 90])

    ws.merge_cells("A1:B1")
    ws["A1"] = "TEMPLATE D'ÉVALUATION COMPARATIVE — SIEM & NAC (3 OFFRES)"
    ws["A1"].fill = FILL_TITLE
    ws["A1"].font = FONT_TITLE
    ws["A1"].alignment = ALIGN_L
    ws.row_dimensions[1].height = 32

    lines = [
        ("", ""),
        ("Objectif", "Comparer objectivement trois offres techniques pour l'acquisition d'un SIEM et d'un NAC, avec notation pondérée et TCO."),
        ("Mode d'emploi", ""),
        ("1", "Remplir la feuille 02_Contexte : votre périmètre, contraintes et pondérations (poids des critères)."),
        ("2", "Identifier les 3 soumissionnaires dans 02_Contexte (Offre A / B / C)."),
        ("3", "Compléter 03_SIEM_Critères et 04_NAC_Critères : réponse Oui/Partiel/Non + note 0–5 + commentaire."),
        ("4", "Saisir les coûts dans 05_TCO (licence, déploiement, support, stockage, formation…)."),
        ("5", "Consulter 06_Scores (calculs automatiques) et 07_Synthese pour la recommandation."),
        ("", ""),
        ("Échelle de notes", ""),
        ("0", "Absent / non conforme"),
        ("1", "Très insuffisant"),
        ("2", "Insuffisant / écarts majeurs"),
        ("3", "Acceptable / écarts mineurs"),
        ("4", "Bon / conforme au besoin"),
        ("5", "Excellent / dépasse le besoin"),
        ("", ""),
        ("Couleurs", ""),
        ("Jaune", "Cellules à remplir (saisie manuelle)"),
        ("Vert", "Cellules calculées automatiquement — ne pas modifier"),
        ("", ""),
        ("Conseil", "Adaptez les poids dans 02_Contexte pour refléter vos priorités (ex. : conformité, TCO, intégration AD, SOC 24/7)."),
        ("", "Conservez une trace des preuves (références pages de l'offre) dans la colonne Commentaire."),
    ]
    r = 3
    for a, b in lines:
        ws.cell(row=r, column=1, value=a).font = FONT_BOLD
        ws.cell(row=r, column=2, value=b).font = FONT_NORMAL
        ws.cell(row=r, column=2).alignment = ALIGN_L
        if a in ("Objectif", "Mode d'emploi", "Échelle de notes", "Couleurs", "Conseil"):
            ws.cell(row=r, column=1).fill = FILL_SECTION
            ws.cell(row=r, column=1).font = FONT_SECTION
            ws.cell(row=r, column=2).fill = FILL_SUB
            ws.cell(row=r, column=2).font = FONT_BOLD
        r += 1

    ws["A20"].fill = FILL_INPUT
    ws["A21"].fill = FILL_CALC


# =============================================================================
# Feuille 2 — Contexte
# =============================================================================
def build_contexte(wb):
    ws = wb.create_sheet("02_Contexte")
    set_col_widths(ws, [36, 28, 28, 28, 40])

    ws.merge_cells("A1:E1")
    ws["A1"] = "CONTEXTE & PÉRIMÈTRE D'ACQUISITION"
    ws["A1"].fill = FILL_TITLE
    ws["A1"].font = FONT_TITLE
    ws.row_dimensions[1].height = 28

    # Identification offres
    style_section(ws, 3, 1, 5, "IDENTIFICATION DES OFFRES")
    headers = ["Champ", "Offre A", "Offre B", "Offre C", "Commentaire"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=4, column=i, value=h)
    style_header_row(ws, 4, 1, 5)

    id_rows = [
        "Nom éditeur / produit SIEM",
        "Nom éditeur / produit NAC",
        "Intégrateur / revendeur",
        "Référence proposition",
        "Date de l'offre",
        "Validité de l'offre",
        "Contact commercial",
        "Modèle de déploiement (On-prem / SaaS / Hybride)",
    ]
    for i, label in enumerate(id_rows):
        r = 5 + i
        ws.cell(row=r, column=1, value=label).font = FONT_BOLD
        ws.cell(row=r, column=1).border = THIN
        for c in range(2, 6):
            style_input(ws.cell(row=r, column=c))

    # Périmètre
    style_section(ws, 14, 1, 5, "PÉRIMÈTRE TECHNIQUE (à renseigner une seule fois)")
    for i, h in enumerate(["Paramètre", "Valeur", "Unité", "Obligation (O/S)", "Notes"], 1):
        ws.cell(row=15, column=i, value=h)
    style_header_row(ws, 15, 1, 5)

    perim = [
        ("Nombre d'utilisateurs / identités", "", "u", "O", ""),
        ("Nombre d'endpoints (EPS) SIEM", "", "EPS", "O", ""),
        ("Volume de logs / jour estimé", "", "Go/jour", "O", ""),
        ("Durée de rétention chaude", "", "jours", "O", ""),
        ("Durée de rétention froide / archive", "", "mois", "S", ""),
        ("Nombre de sites / datacenters", "", "sites", "O", ""),
        ("Nombre de switches / contrôleurs Wi-Fi (NAC)", "", "équip.", "O", ""),
        ("Couverture Wi-Fi / filaire / VPN / IoT", "", "texte", "O", ""),
        ("Intégration AD / Entra ID / LDAP", "", "Oui/Non", "O", ""),
        ("Exigence Haute Disponibilité", "", "Oui/Non", "S", ""),
        ("Exigence hébergement données (UE / souveraineté)", "", "texte", "O", ""),
        ("SOC interne / MSSP / hybride", "", "texte", "O", ""),
        ("Budget max TCO 3 ans (indicatif)", "", "€ HT", "S", ""),
        ("Date cible de mise en production", "", "date", "S", ""),
    ]
    for i, (p, v, u, o, n) in enumerate(perim):
        r = 16 + i
        ws.cell(row=r, column=1, value=p).font = FONT_NORMAL
        ws.cell(row=r, column=1).border = THIN
        style_input(ws.cell(row=r, column=2, value=v))
        ws.cell(row=r, column=3, value=u).border = THIN
        style_input(ws.cell(row=r, column=4, value=o))
        style_input(ws.cell(row=r, column=5, value=n))

    # Pondérations SIEM
    style_section(ws, 32, 1, 5, "PONDÉRATION DES CRITÈRES SIEM (total = 100 %)")
    for i, h in enumerate(["Famille de critères", "Poids %", "Offre A (auto)", "Offre B (auto)", "Offre C (auto)"], 1):
        ws.cell(row=33, column=i, value=h)
    style_header_row(ws, 33, 1, 5)

    siem_weights = [
        ("Fonctionnalités & détection", 25),
        ("Architecture & performance", 15),
        ("Intégrations & écosystème", 15),
        ("Exploitation & usabilité SOC", 15),
        ("Sécurité, conformité, gouvernance", 10),
        ("Support, formation, roadmap", 10),
        ("TCO / rapport qualité-prix", 10),
    ]
    for i, (name, w) in enumerate(siem_weights):
        r = 34 + i
        ws.cell(row=r, column=1, value=name).border = THIN
        style_input(ws.cell(row=r, column=2, value=w))
        for c in range(3, 6):
            style_calc(ws.cell(row=r, column=c))
            ws.cell(row=r, column=c).value = "→ voir 06_Scores"

    ws.cell(row=41, column=1, value="TOTAL poids SIEM").font = FONT_BOLD
    style_calc(ws.cell(row=41, column=2, value="=SUM(B34:B40)"))
    ws["B42"] = "Le total doit être égal à 100"
    ws["B42"].font = FONT_SMALL

    # Pondérations NAC
    style_section(ws, 44, 1, 5, "PONDÉRATION DES CRITÈRES NAC (total = 100 %)")
    for i, h in enumerate(["Famille de critères", "Poids %", "Offre A (auto)", "Offre B (auto)", "Offre C (auto)"], 1):
        ws.cell(row=45, column=i, value=h)
    style_header_row(ws, 45, 1, 5)

    nac_weights = [
        ("Contrôle d'accès & politiques", 25),
        ("Profilage & visibilité assets", 15),
        ("Intégrations réseau / IAM / MDM", 20),
        ("Déploiement & modes (monitor→enforce)", 10),
        ("Sécurité, conformité, segmentation", 10),
        ("Support, formation, roadmap", 10),
        ("TCO / rapport qualité-prix", 10),
    ]
    for i, (name, w) in enumerate(nac_weights):
        r = 46 + i
        ws.cell(row=r, column=1, value=name).border = THIN
        style_input(ws.cell(row=r, column=2, value=w))
        for c in range(3, 6):
            style_calc(ws.cell(row=r, column=c))
            ws.cell(row=r, column=c).value = "→ voir 06_Scores"

    ws.cell(row=53, column=1, value="TOTAL poids NAC").font = FONT_BOLD
    style_calc(ws.cell(row=53, column=2, value="=SUM(B46:B52)"))
    ws["B54"] = "Le total doit être égal à 100"
    ws["B54"].font = FONT_SMALL

    # Pondération globale SIEM vs NAC
    style_section(ws, 56, 1, 5, "PONDÉRATION GLOBALE PROJET")
    ws["A57"] = "Poids SIEM dans le score global (%)"
    style_input(ws.cell(row=57, column=2, value=55))
    ws["A58"] = "Poids NAC dans le score global (%)"
    style_input(ws.cell(row=58, column=2, value=45))
    ws["A59"] = "TOTAL"
    style_calc(ws.cell(row=59, column=2, value="=B57+B58"))
    ws["C57"] = "Doit totaliser 100 %"
    ws["C57"].font = FONT_SMALL


# =============================================================================
# Critères SIEM / NAC
# =============================================================================
SIEM_CRITERIA = [
    ("FONCTIONNALITÉS & DÉTECTION", None),
    ("Collecte multi-sources (syslog, agents, API, cloud)", "Fonctionnalités & détection"),
    ("Corrélation temps réel / règles custom", "Fonctionnalités & détection"),
    ("UEBA / détection comportementale / ML", "Fonctionnalités & détection"),
    ("Threat intelligence intégrée (feeds)", "Fonctionnalités & détection"),
    ("SOAR / playbooks / automatisation", "Fonctionnalités & détection"),
    ("Cas d'usage métiers pré-packagés", "Fonctionnalités & détection"),
    ("Recherche forensics / hunting", "Fonctionnalités & détection"),
    ("Tableaux de bord & reporting (KPI, conformité)", "Fonctionnalités & détection"),
    ("ARCHITECTURE & PERFORMANCE", None),
    ("Scalabilité (EPS / volume)", "Architecture & performance"),
    ("Haute disponibilité / PRA-PCA", "Architecture & performance"),
    ("Rétention chaude / froide / archive", "Architecture & performance"),
    ("Latence d'ingestion & alerting", "Architecture & performance"),
    ("Architecture On-prem / SaaS / hybride proposée", "Architecture & performance"),
    ("INTÉGRATIONS & ÉCOSYSTÈME", None),
    ("Intégration AD / Entra ID / IAM", "Intégrations & écosystème"),
    ("Intégration EDR / XDR / antivirus", "Intégrations & écosystème"),
    ("Intégration firewall / proxy / mail gateway", "Intégrations & écosystème"),
    ("Intégration ticketing (ServiceNow, Jira…)", "Intégrations & écosystème"),
    ("Intégration avec le NAC (événements / politiques)", "Intégrations & écosystème"),
    ("Connecteurs cloud (M365, AWS, Azure, GCP)", "Intégrations & écosystème"),
    ("EXPLOITATION & USABILITÉ SOC", None),
    ("Ergonomie console / courbe d'apprentissage", "Exploitation & usabilité SOC"),
    ("Gestion des faux positifs / tuning", "Exploitation & usabilité SOC"),
    ("RBAC / multi-tenancy / équipes", "Exploitation & usabilité SOC"),
    ("Charge d'administration estimée", "Exploitation & usabilité SOC"),
    ("Documentation & contenu (règles, parsers)", "Exploitation & usabilité SOC"),
    ("SÉCURITÉ, CONFORMITÉ, GOUVERNANCE", None),
    ("Chiffrement données au repos / en transit", "Sécurité, conformité, gouvernance"),
    ("Hébergement / souveraineté des données", "Sécurité, conformité, gouvernance"),
    ("Traçabilité admin / audit trail", "Sécurité, conformité, gouvernance"),
    ("Conformité (RGPD, ISO 27001, NIS2, secteur)", "Sécurité, conformité, gouvernance"),
    ("SUPPORT, FORMATION, ROADMAP", None),
    ("SLA support (horaires, sévérités)", "Support, formation, roadmap"),
    ("Formation incluse / transfert de compétences", "Support, formation, roadmap"),
    ("Roadmap produit & pérennité éditeur", "Support, formation, roadmap"),
    ("Références clients sectorielles", "Support, formation, roadmap"),
    ("TCO / RAPPORT QUALITÉ-PRIX", None),
    ("Clarté du modèle de licence", "TCO / rapport qualité-prix"),
    ("Adéquation prix / périmètre", "TCO / rapport qualité-prix"),
    ("Coûts cachés (stockage, connecteurs, EPS burst)", "TCO / rapport qualité-prix"),
]

NAC_CRITERIA = [
    ("CONTRÔLE D'ACCÈS & POLITIQUES", None),
    ("802.1X filaire / Wi-Fi", "Contrôle d'accès & politiques"),
    ("MAB / MAC authentication bypass", "Contrôle d'accès & politiques"),
    ("Portail captive / accès invités", "Contrôle d'accès & politiques"),
    ("Politiques dynamiques (VLAN, ACL, Quarantine)", "Contrôle d'accès & politiques"),
    ("Posture assessment (antivirus, patch, disk encrypt)", "Contrôle d'accès & politiques"),
    ("Gestion BYOD / employés / partenaires", "Contrôle d'accès & politiques"),
    ("Gestion IoT / OT / équipements non managés", "Contrôle d'accès & politiques"),
    ("PROFILAGE & VISIBILITÉ", None),
    ("Découverte & inventaire des assets", "Profilage & visibilité assets"),
    ("Fingerprint / profiling passif & actif", "Profilage & visibilité assets"),
    ("Classification automatique des devices", "Profilage & visibilité assets"),
    ("Visibilité temps réel (qui/quoi/où)", "Profilage & visibilité assets"),
    ("INTÉGRATIONS RÉSEAU / IAM / MDM", None),
    ("Compatibilité switches multi-vendeurs", "Intégrations réseau / IAM / MDM"),
    ("Contrôleurs Wi-Fi / WLC", "Intégrations réseau / IAM / MDM"),
    ("VPN / remote access", "Intégrations réseau / IAM / MDM"),
    ("Intégration AD / Entra ID / RADIUS / PKI", "Intégrations réseau / IAM / MDM"),
    ("Intégration MDM / EMM (Intune, etc.)", "Intégrations réseau / IAM / MDM"),
    ("Intégration SIEM / firewall / SOAR", "Intégrations réseau / IAM / MDM"),
    ("DÉPLOIEMENT & MODES", None),
    ("Mode monitor / audit puis enforce", "Déploiement & modes (monitor→enforce)"),
    ("Complexité de déploiement (agents, appliances)", "Déploiement & modes (monitor→enforce)"),
    ("Haute disponibilité du policy server", "Déploiement & modes (monitor→enforce)"),
    ("Scalabilité multi-sites", "Déploiement & modes (monitor→enforce)"),
    ("SÉCURITÉ & SEGMENTATION", None),
    ("Micro-segmentation / Zero Trust network access", "Sécurité, conformité, segmentation"),
    ("Quarantine / remédiation automatisée", "Sécurité, conformité, segmentation"),
    ("Traçabilité & reporting conformité", "Sécurité, conformité, segmentation"),
    ("SUPPORT, FORMATION, ROADMAP", None),
    ("SLA support", "Support, formation, roadmap"),
    ("Formation & documentation", "Support, formation, roadmap"),
    ("Roadmap & pérennité", "Support, formation, roadmap"),
    ("Références clients", "Support, formation, roadmap"),
    ("TCO / RAPPORT QUALITÉ-PRIX", None),
    ("Clarté du modèle de licence", "TCO / rapport qualité-prix"),
    ("Adéquation prix / périmètre", "TCO / rapport qualité-prix"),
    ("Coûts appliances / licences endpoints / renew", "TCO / rapport qualité-prix"),
]


def build_criteria_sheet(wb, title, criteria, product_label):
    ws = wb.create_sheet(title)
    # Colonnes: Critère | Famille | A Conformité | A Note | A Commentaire | B ... | C ... | Preuve / page
    headers = [
        "Critère",
        "Famille",
        "A — Conformité",
        "A — Note (0-5)",
        "A — Commentaire",
        "B — Conformité",
        "B — Note (0-5)",
        "B — Commentaire",
        "C — Conformité",
        "C — Note (0-5)",
        "C — Commentaire",
        "Preuve / réf. offre",
    ]
    set_col_widths(ws, [48, 28, 14, 12, 28, 14, 12, 28, 14, 12, 28, 22])

    ws.merge_cells("A1:L1")
    ws["A1"] = f"GRILLE D'ÉVALUATION TECHNIQUE — {product_label}"
    ws["A1"].fill = FILL_TITLE
    ws["A1"].font = FONT_TITLE
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:L2")
    ws["A2"] = (
        "Remplir les colonnes jaunes. Conformité = Oui / Partiel / Non / N/A. "
        "Note = 0 (absent) à 5 (excellent). Les totaux sont dans 06_Scores."
    )
    ws["A2"].font = FONT_SMALL

    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h)
    style_header_row(ws, 3, 1, 12)
    ws.freeze_panes = "C4"
    ws.auto_filter.ref = "A3:L3"

    score_cells_a, score_cells_b, score_cells_c = [], [], []
    conf_ranges = []

    r = 4
    for label, family in criteria:
        if family is None:
            style_section(ws, r, 1, 12, label)
            r += 1
            continue
        ws.cell(row=r, column=1, value=label).font = FONT_NORMAL
        ws.cell(row=r, column=1).alignment = ALIGN_L
        ws.cell(row=r, column=1).border = THIN
        ws.cell(row=r, column=2, value=family).border = THIN
        ws.cell(row=r, column=2).fill = FILL_SUB

        for c in (3, 4, 5, 6, 7, 8, 9, 10, 11, 12):
            style_input(ws.cell(row=r, column=c))

        score_cells_a.append(f"D{r}")
        score_cells_b.append(f"G{r}")
        score_cells_c.append(f"J{r}")
        conf_ranges.append(f"C{r}")
        conf_ranges.append(f"F{r}")
        conf_ranges.append(f"I{r}")
        r += 1

    last_data = r - 1
    if score_cells_a:
        add_score_validation(
            ws,
            f"D4:D{last_data}",
            f"G4:G{last_data}",
            f"J4:J{last_data}",
        )
        add_oui_non(
            ws,
            f"C4:C{last_data}",
            f"F4:F{last_data}",
            f"I4:I{last_data}",
        )

        # Color scale on notes
        for col in ("D", "G", "J"):
            ws.conditional_formatting.add(
                f"{col}4:{col}{last_data}",
                ColorScaleRule(
                    start_type="num", start_value=0, start_color="F8696B",
                    mid_type="num", mid_value=3, mid_color="FFEB84",
                    end_type="num", end_value=5, end_color="63BE7B",
                ),
            )

    # Ligne moyenne simple
    ws.cell(row=r, column=1, value="MOYENNE SIMPLE DES NOTES (indicatif)").font = FONT_BOLD
    style_calc(ws.cell(row=r, column=4, value=f'=IFERROR(AVERAGE(D4:D{last_data}),"")'))
    style_calc(ws.cell(row=r, column=7, value=f'=IFERROR(AVERAGE(G4:G{last_data}),"")'))
    style_calc(ws.cell(row=r, column=10, value=f'=IFERROR(AVERAGE(J4:J{last_data}),"")'))

    return ws


# =============================================================================
# TCO
# =============================================================================
def build_tco(wb):
    ws = wb.create_sheet("05_TCO")
    set_col_widths(ws, [42, 16, 16, 16, 16, 28])

    ws.merge_cells("A1:F1")
    ws["A1"] = "COÛT TOTAL DE POSSESSION (TCO) — 3 ANS — € HT"
    ws["A1"].fill = FILL_TITLE
    ws["A1"].font = FONT_TITLE
    ws.row_dimensions[1].height = 28

    headers = ["Poste de coût", "Offre A", "Offre B", "Offre C", "Unité / base", "Commentaire / hypothèses"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h)
    style_header_row(ws, 3, 1, 6)

    sections = [
        ("SIEM — LICENCES & ABONNEMENTS", [
            "Licence / abonnement SIEM année 1",
            "Licence / abonnement SIEM année 2",
            "Licence / abonnement SIEM année 3",
            "Modules optionnels (UEBA, SOAR, TI…)",
            "Stockage / rétention additionnelle",
            "Connecteurs / EPS burst / overage",
        ]),
        ("NAC — LICENCES & ABONNEMENTS", [
            "Licence / abonnement NAC année 1",
            "Licence / abonnement NAC année 2",
            "Licence / abonnement NAC année 3",
            "Licences endpoints / devices",
            "Appliances / VM / hardware",
            "Modules optionnels (guest, posture…)",
        ]),
        ("DÉPLOIEMENT & INTÉGRATION", [
            "Prestation d'intégration SIEM",
            "Prestation d'intégration NAC",
            "Développement règles / cas d'usage",
            "Migration / reprise de l'existant",
            "Frais de projet / PMO",
        ]),
        ("FORMATION & ACCOMPAGNEMENT", [
            "Formation administrateurs",
            "Formation analystes SOC",
            "Run / assistance au démarrage (jours)",
        ]),
        ("SUPPORT & MAINTENANCE", [
            "Support / maintenance année 1 (si séparé)",
            "Support / maintenance année 2",
            "Support / maintenance année 3",
            "MSSP / co-gestion SOC (option)",
        ]),
        ("AUTRES", [
            "Infrastructure d'hébergement (si on-prem)",
            "Réseau / certificats / PKI",
            "Divers / imprévus (ex. 5–10 %)",
        ]),
    ]

    r = 4
    money_rows = []
    for section, items in sections:
        style_section(ws, r, 1, 6, section)
        r += 1
        for item in items:
            ws.cell(row=r, column=1, value=item).border = THIN
            for c in range(2, 5):
                style_input(ws.cell(row=r, column=c, value=0))
                ws.cell(row=r, column=c).number_format = '#,##0 €'
            style_input(ws.cell(row=r, column=5))
            style_input(ws.cell(row=r, column=6))
            money_rows.append(r)
            r += 1

    # Totaux
    style_section(ws, r, 1, 6, "TOTAUX")
    r += 1
    first, last = money_rows[0], money_rows[-1]
    for label, formula_tpl in [
        ("SOUS-TOTAL SIEM (licences liées)", None),
    ]:
        pass

    ws.cell(row=r, column=1, value="TCO 3 ANS TOTAL").font = FONT_BOLD
    for c, col in enumerate(("B", "C", "D"), 2):
        style_calc(ws.cell(row=r, column=c, value=f"=SUM({col}{first}:{col}{last})"))
        ws.cell(row=r, column=c).number_format = '#,##0 €'
    total_row = r
    r += 2

    ws.cell(row=r, column=1, value="TCO annuel moyen").font = FONT_BOLD
    for c, col in enumerate(("B", "C", "D"), 2):
        style_calc(ws.cell(row=r, column=c, value=f"={col}{total_row}/3"))
        ws.cell(row=r, column=c).number_format = '#,##0 €'
    r += 2

    ws.cell(row=r, column=1, value="Note TCO (0–5) — à reporter dans les grilles si besoin").font = FONT_BOLD
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=1)
    r += 1
    ws.cell(row=r, column=1, value="Plus bas TCO = meilleure note (saisie manuelle ou auto ci-dessous)")
    ws["A" + str(r)].font = FONT_SMALL
    r += 1
    # Auto note: 5 for cheapest, proportional
    # We'll put formulas that score relative to min
    note_row = r
    ws.cell(row=r, column=1, value="Note auto TCO (5 = moins cher)").font = FONT_BOLD
    # Note = 5 * min / value (capped) — if value 0, blank
    for c, col in enumerate(("B", "C", "D"), 2):
        style_calc(
            ws.cell(
                row=r,
                column=c,
                value=f'=IF({col}{total_row}=0,"",ROUND(5*MIN($B${total_row}:$D${total_row})/{col}{total_row},2))',
            )
        )
    r += 2
    ws.cell(row=r, column=1, value="Hypothèses globales (inflation, taux de change, durée contrat…)").font = FONT_BOLD
    style_input(ws.cell(row=r + 1, column=1))
    ws.merge_cells(start_row=r + 1, start_column=1, end_row=r + 3, end_column=6)
    for rr in range(r + 1, r + 4):
        for c in range(1, 7):
            style_input(ws.cell(row=rr, column=c))

    return total_row, note_row


# =============================================================================
# Scores
# =============================================================================
def build_scores(wb):
    ws = wb.create_sheet("06_Scores")
    set_col_widths(ws, [40, 14, 14, 14, 18, 18, 18, 22])

    ws.merge_cells("A1:H1")
    ws["A1"] = "SCORES PONDÉRÉS — CALCUL AUTOMATIQUE"
    ws["A1"].fill = FILL_TITLE
    ws["A1"].font = FONT_TITLE
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:H2")
    ws["A2"] = (
        "Les moyennes par famille sont calculées depuis 03_SIEM_Critères et 04_NAC_Critères. "
        "Les poids viennent de 02_Contexte. Ne modifiez que les poids / notes sources."
    )
    ws["A2"].font = FONT_SMALL

    # --- SIEM ---
    style_section(ws, 4, 1, 8, "SIEM — SCORES PAR FAMILLE")
    headers = [
        "Famille",
        "Poids %",
        "Moy. A",
        "Moy. B",
        "Moy. C",
        "Score pond. A",
        "Score pond. B",
        "Score pond. C",
    ]
    for i, h in enumerate(headers, 1):
        ws.cell(row=5, column=i, value=h)
    style_header_row(ws, 5, 1, 8)

    siem_families = [
        ("Fonctionnalités & détection", "'02_Contexte'!B34"),
        ("Architecture & performance", "'02_Contexte'!B35"),
        ("Intégrations & écosystème", "'02_Contexte'!B36"),
        ("Exploitation & usabilité SOC", "'02_Contexte'!B37"),
        ("Sécurité, conformité, gouvernance", "'02_Contexte'!B38"),
        ("Support, formation, roadmap", "'02_Contexte'!B39"),
        ("TCO / rapport qualité-prix", "'02_Contexte'!B40"),
    ]

    # AVERAGEIF on family column B of criteria sheet, scores in D/G/J
    for i, (fam, weight_ref) in enumerate(siem_families):
        r = 6 + i
        ws.cell(row=r, column=1, value=fam).border = THIN
        style_calc(ws.cell(row=r, column=2, value=f"={weight_ref}"))
        # Averageif
        for col_idx, score_col in [(3, "D"), (4, "G"), (5, "J")]:
            style_calc(
                ws.cell(
                    row=r,
                    column=col_idx,
                    value=f"=IFERROR(AVERAGEIF('03_SIEM_Critères'!$B:$B,A{r},'03_SIEM_Critères'!{score_col}:{score_col}),\"\")",
                )
            )
        for col_idx, moy_col in [(6, "C"), (7, "D"), (8, "E")]:
            style_calc(
                ws.cell(
                    row=r,
                    column=col_idx,
                    value=f'=IF({moy_col}{r}="","",{moy_col}{r}*B{r}/100)',
                )
            )

    r = 13
    ws.cell(row=r, column=1, value="SCORE SIEM / 5").font = FONT_BOLD
    style_calc(ws.cell(row=r, column=2, value="=SUM(B6:B12)"))
    for c, col in [(6, "F"), (7, "G"), (8, "H")]:
        style_calc(ws.cell(row=r, column=c, value=f"=SUM({col}6:{col}12)"))
    siem_total_row = r

    # --- NAC ---
    style_section(ws, 15, 1, 8, "NAC — SCORES PAR FAMILLE")
    for i, h in enumerate(headers, 1):
        ws.cell(row=16, column=i, value=h)
    style_header_row(ws, 16, 1, 8)

    nac_families = [
        ("Contrôle d'accès & politiques", "'02_Contexte'!B46"),
        ("Profilage & visibilité assets", "'02_Contexte'!B47"),
        ("Intégrations réseau / IAM / MDM", "'02_Contexte'!B48"),
        ("Déploiement & modes (monitor→enforce)", "'02_Contexte'!B49"),
        ("Sécurité, conformité, segmentation", "'02_Contexte'!B50"),
        ("Support, formation, roadmap", "'02_Contexte'!B51"),
        ("TCO / rapport qualité-prix", "'02_Contexte'!B52"),
    ]

    for i, (fam, weight_ref) in enumerate(nac_families):
        r = 17 + i
        ws.cell(row=r, column=1, value=fam).border = THIN
        style_calc(ws.cell(row=r, column=2, value=f"={weight_ref}"))
        for col_idx, score_col in [(3, "D"), (4, "G"), (5, "J")]:
            style_calc(
                ws.cell(
                    row=r,
                    column=col_idx,
                    value=f"=IFERROR(AVERAGEIF('04_NAC_Critères'!$B:$B,A{r},'04_NAC_Critères'!{score_col}:{score_col}),\"\")",
                )
            )
        for col_idx, moy_col in [(6, "C"), (7, "D"), (8, "E")]:
            style_calc(
                ws.cell(
                    row=r,
                    column=col_idx,
                    value=f'=IF({moy_col}{r}="","",{moy_col}{r}*B{r}/100)',
                )
            )

    r = 24
    ws.cell(row=r, column=1, value="SCORE NAC / 5").font = FONT_BOLD
    style_calc(ws.cell(row=r, column=2, value="=SUM(B17:B23)"))
    for c, col in [(6, "F"), (7, "G"), (8, "H")]:
        style_calc(ws.cell(row=r, column=c, value=f"=SUM({col}17:{col}23)"))
    nac_total_row = r

    # --- Global ---
    style_section(ws, 26, 1, 8, "SCORE GLOBAL PROJET (SIEM + NAC)")
    for i, h in enumerate(["Composante", "Poids %", "", "", "", "Offre A", "Offre B", "Offre C"], 1):
        ws.cell(row=27, column=i, value=h)
    style_header_row(ws, 27, 1, 8)

    ws["A28"] = "SIEM"
    style_calc(ws.cell(row=28, column=2, value="='02_Contexte'!B57"))
    style_calc(ws.cell(row=28, column=6, value=f"=F{siem_total_row}"))
    style_calc(ws.cell(row=28, column=7, value=f"=G{siem_total_row}"))
    style_calc(ws.cell(row=28, column=8, value=f"=H{siem_total_row}"))

    ws["A29"] = "NAC"
    style_calc(ws.cell(row=29, column=2, value="='02_Contexte'!B58"))
    style_calc(ws.cell(row=29, column=6, value=f"=F{nac_total_row}"))
    style_calc(ws.cell(row=29, column=7, value=f"=G{nac_total_row}"))
    style_calc(ws.cell(row=29, column=8, value=f"=H{nac_total_row}"))

    ws["A30"] = "SCORE GLOBAL / 5"
    ws["A30"].font = FONT_BOLD
    for col in ("F", "G", "H"):
        style_calc(
            ws.cell(
                row=30,
                column=ord(col) - ord("A") + 1,
                value=f"=({col}28*B28+{col}29*B29)/100",
            )
        )

    ws["A32"] = "Classement (1 = meilleur)"
    ws["A32"].font = FONT_BOLD
    # Rank: higher score better
    for col, cell in [("F", "F33"), ("G", "G33"), ("H", "H33")]:
        pass
    style_calc(ws.cell(row=33, column=6, value='=IF(F30="","",RANK(F30,$F$30:$H$30,0))'))
    style_calc(ws.cell(row=33, column=7, value='=IF(G30="","",RANK(G30,$F$30:$H$30,0))'))
    style_calc(ws.cell(row=33, column=8, value='=IF(H30="","",RANK(H30,$F$30:$H$30,0))'))
    ws["A33"] = "Rang"
    ws["F32"] = "Offre A"
    ws["G32"] = "Offre B"
    ws["H32"] = "Offre C"
    for c in range(6, 9):
        ws.cell(row=32, column=c).font = FONT_BOLD
        ws.cell(row=32, column=c).alignment = ALIGN_C

    # Color scale on global scores
    ws.conditional_formatting.add(
        "F30:H30",
        ColorScaleRule(
            start_type="min", start_color="F8696B",
            mid_type="percentile", mid_value=50, mid_color="FFEB84",
            end_type="max", end_color="63BE7B",
        ),
    )


# =============================================================================
# Synthèse
# =============================================================================
def build_synthese(wb):
    ws = wb.create_sheet("07_Synthese")
    set_col_widths(ws, [36, 22, 22, 22, 40])

    ws.merge_cells("A1:E1")
    ws["A1"] = "SYNTHÈSE DÉCISIONNELLE"
    ws["A1"].fill = FILL_TITLE
    ws["A1"].font = FONT_TITLE
    ws.row_dimensions[1].height = 28

    style_section(ws, 3, 1, 5, "TABLEAU DE BORD")
    headers = ["Indicateur", "Offre A", "Offre B", "Offre C", "Commentaire"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=4, column=i, value=h)
    style_header_row(ws, 4, 1, 5)

    rows = [
        ("Éditeur / produit SIEM", "='02_Contexte'!B5", "='02_Contexte'!C5", "='02_Contexte'!D5"),
        ("Éditeur / produit NAC", "='02_Contexte'!B6", "='02_Contexte'!C6", "='02_Contexte'!D6"),
        ("Score SIEM / 5", "='06_Scores'!F13", "='06_Scores'!G13", "='06_Scores'!H13"),
        ("Score NAC / 5", "='06_Scores'!F24", "='06_Scores'!G24", "='06_Scores'!H24"),
        ("Score global / 5", "='06_Scores'!F30", "='06_Scores'!G30", "='06_Scores'!H30"),
        ("Classement", "='06_Scores'!F33", "='06_Scores'!G33", "='06_Scores'!H33"),
        ("TCO 3 ans (€ HT)", "='05_TCO'!B35", "='05_TCO'!C35", "='05_TCO'!D35"),
    ]
    # TCO row number may vary — we'll fix by named approach: use a safer formula later
    # Actually TCO total row is dynamic. Let me recalculate - better to use a fixed known structure.
    # Looking at build_tco: sections have known item counts...
    # I'll put placeholder formulas and fix TCO row after building by searching.
    # For now use a note that TCO links from 05_TCO total row — I'll patch after.

    for i, (label, a, b, c) in enumerate(rows):
        r = 5 + i
        ws.cell(row=r, column=1, value=label).font = FONT_BOLD
        ws.cell(row=r, column=1).border = THIN
        for col, val in [(2, a), (3, b), (4, c)]:
            style_calc(ws.cell(row=r, column=col, value=val))
        style_input(ws.cell(row=r, column=5))

    # Forces / faiblesses
    style_section(ws, 14, 1, 5, "ANALYSE QUALITATIVE (saisie libre)")
    for i, h in enumerate(["Offre", "Points forts", "Points faibles / écarts", "Risques", "Conditions suspensives"], 1):
        ws.cell(row=15, column=i, value=h)
    style_header_row(ws, 15, 1, 5)

    for i, name in enumerate(["Offre A", "Offre B", "Offre C"]):
        r = 16 + i
        ws.cell(row=r, column=1, value=name).font = FONT_BOLD
        ws.cell(row=r, column=1).border = THIN
        ws.row_dimensions[r].height = 60
        for c in range(2, 6):
            style_input(ws.cell(row=r, column=c))
            ws.cell(row=r, column=c).alignment = ALIGN_L

    style_section(ws, 20, 1, 5, "RECOMMANDATION")
    ws["A21"] = "Offre recommandée"
    ws["A21"].font = FONT_BOLD
    style_input(ws.cell(row=21, column=2))
    ws.merge_cells("B21:C21")

    ws["A22"] = "Justification (5–10 lignes)"
    ws["A22"].font = FONT_BOLD
    ws.merge_cells("A23:E26")
    for r in range(23, 27):
        for c in range(1, 6):
            style_input(ws.cell(row=r, column=c))
    ws["A23"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    ws["A28"] = "Scénario alternatif / plan B"
    ws["A28"].font = FONT_BOLD
    ws.merge_cells("A29:E30")
    for r in range(29, 31):
        for c in range(1, 6):
            style_input(ws.cell(row=r, column=c))

    ws["A32"] = "Prochaines étapes (PoC, audit licence, négociation, atelier technique…)"
    ws["A32"].font = FONT_BOLD
    ws.merge_cells("A33:E35")
    for r in range(33, 36):
        for c in range(1, 6):
            style_input(ws.cell(row=r, column=c))

    ws["A37"] = "Validation"
    ws["A37"].fill = FILL_SECTION
    ws["A37"].font = FONT_SECTION
    style_section(ws, 37, 1, 5, "VALIDATION")
    for i, h in enumerate(["Rôle", "Nom", "Date", "Visa / décision", "Commentaire"], 1):
        ws.cell(row=38, column=i, value=h)
    style_header_row(ws, 38, 1, 5)
    for role in ["RSSI / Cyber", "DSI / Infra", "Achats", "Direction"]:
        r = 39 + ["RSSI / Cyber", "DSI / Infra", "Achats", "Direction"].index(role)
        ws.cell(row=r, column=1, value=role).border = THIN
        for c in range(2, 6):
            style_input(ws.cell(row=r, column=c))


def fix_tco_links(wb, tco_total_row):
    """Corrige les formules TCO dans la synthèse une fois le n° de ligne connu."""
    ws = wb["07_Synthese"]
    # Row 11 = TCO 3 ans (5+6 = index 6 in rows → row 11)
    ws["B11"] = f"='05_TCO'!B{tco_total_row}"
    ws["C11"] = f"='05_TCO'!C{tco_total_row}"
    ws["D11"] = f"='05_TCO'!D{tco_total_row}"
    for col in ("B", "C", "D"):
        ws[f"{col}11"].number_format = '#,##0 €'
        style_calc(ws[f"{col}11"])


def main():
    wb = Workbook()
    build_instructions(wb)
    build_contexte(wb)
    build_criteria_sheet(wb, "03_SIEM_Critères", SIEM_CRITERIA, "SIEM")
    build_criteria_sheet(wb, "04_NAC_Critères", NAC_CRITERIA, "NAC")
    tco_total_row, _ = build_tco(wb)
    build_scores(wb)
    build_synthese(wb)
    fix_tco_links(wb, tco_total_row)

    # Feuille checklist PoC optionnelle
    ws = wb.create_sheet("08_Checklist_PoC")
    set_col_widths(ws, [8, 50, 14, 14, 14, 30])
    ws.merge_cells("A1:F1")
    ws["A1"] = "CHECKLIST PoC / PILOTE (optionnel)"
    ws["A1"].fill = FILL_TITLE
    ws["A1"].font = FONT_TITLE

    headers = ["#", "Scénario de test", "Offre A", "Offre B", "Offre C", "Résultat / écart"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h)
    style_header_row(ws, 3, 1, 6)

    poc_items = [
        "Ingestion des sources de logs prioritaires",
        "Création / tuning d'un cas d'usage de détection",
        "Temps de recherche forensics sur 7 jours",
        "Génération d'un rapport conformité",
        "Intégration EDR → alerte SIEM",
        "Authentification 802.1X sur switch test",
        "Profilage d'un device IoT inconnu",
        "Mise en quarantine automatisée",
        "Portail invité + sponsoring",
        "Intégration AD / certificats",
        "Corrélation événement NAC → SIEM",
        "Basculer monitor → enforce sans coupure",
        "Failover HA (si exigé)",
        "Export / API / ticketing",
        "Charge / performance sous volume cible",
    ]
    add_oui_non(ws, "C4:E18")
    for i, item in enumerate(poc_items, 1):
        r = 3 + i
        ws.cell(row=r, column=1, value=i).border = THIN
        ws.cell(row=r, column=1).alignment = ALIGN_C
        ws.cell(row=r, column=2, value=item).border = THIN
        for c in range(3, 7):
            style_input(ws.cell(row=r, column=c))

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Template généré : {OUT}")


if __name__ == "__main__":
    main()
